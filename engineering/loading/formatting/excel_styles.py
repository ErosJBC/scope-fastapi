"""
A module for Excel styles in the engineering.loading.formatting package.
"""

import pandas as pd
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from engineering.loading.formatting.utils import get_excel_column_letter, get_cols_widths


def generate_styles() -> dict[str, Any]:
    """
    Generate styles for the Excel file.

    :return: A dictionary of styles
    :rtype: dict[str, Any]
    """
    bold_white_font = Font(bold=True, color='FFFFFF')
    bold_black_font = Font(bold=True, color='000000')
    dark_blue_fill = PatternFill(start_color='0F243E', end_color='0F243E', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    top_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    dict_styles: dict[str, Any] = {
        'bold_white_font': bold_white_font,
        'bold_black_font': bold_black_font,
        'dark_blue_fill': dark_blue_fill,
        'white_fill': white_fill,
        'yellow_fill': yellow_fill,
        'top_alignment': top_alignment,
        'thin_border': thin_border,
        'number': '0',
        'numeric': '#,##0.00',
        'percentage': '0.00%',
    }

    return dict_styles

def subtotal_format(data: pd.DataFrame, worksheet: Worksheet, list_columns: list[str]) -> None:
    """
    Apply formatting to the subtotal of the worksheet.

    :param data: The exported dataframe
    :type data: pd.DataFrame
    :param worksheet: The worksheet to apply formatting to
    :type worksheet: Worksheet
    :param list_columns: The list of columns to apply the subtotal
    :type list_columns: list[str]
    """
    for column_num, column_name in enumerate(data.columns):
        column_letter: str = get_excel_column_letter(column_num + 1)
        if column_name in list_columns:
            formula = f'=SUBTOTAL(9,{column_letter}3:{column_letter}{len(data) + 2})'
            worksheet[f'{column_letter}1'] = formula
            # Add formatting to the subtotal
            worksheet[f'{column_letter}1'].font = styles["bold_black_font"]
            worksheet[f'{column_letter}1'].fill = styles["yellow_fill"]
            worksheet[f'{column_letter}1'].border = styles["thin_border"]
            worksheet[f'{column_letter}1'].number_format = styles["numeric"]

def header_format(data: pd.DataFrame, worksheet: Worksheet) -> None:
    """
    Apply formatting to the header of the worksheet.

    :param data: The exported dataframe
    :type data: pd.DataFrame
    :param worksheet: The worksheet to apply formatting to
    :type worksheet: Worksheet
    """
    for column_num, (value, width) in enumerate(zip(data.columns.values, get_cols_widths(data))):
        worksheet.column_dimensions[get_excel_column_letter(column_num + 1)].width = width + 2
        worksheet[f'{get_excel_column_letter(column_num + 1)}2'] = value
        # Add formatting to the header
        worksheet[f'{get_excel_column_letter(column_num + 1)}2'].font = styles["bold_white_font"]
        worksheet[f'{get_excel_column_letter(column_num + 1)}2'].fill = styles["dark_blue_fill"]
        worksheet[f'{get_excel_column_letter(column_num + 1)}2'].alignment = styles["top_alignment"]

def summary_format(data: pd.DataFrame, worksheet: Worksheet) -> None:
    """
    Load dataframes into an Excel file and apply formatting.

    :param data: The exported dataframe
    :type data: pd.DataFrame
    :param worksheet: The worksheet to apply formatting to
    :type worksheet: Worksheet
    """
    for column_num, value in enumerate(data.columns.values):
        worksheet[f'{get_excel_column_letter(column_num + 1)}1'] = value
        # Add formatting to the header
        worksheet[f'{get_excel_column_letter(column_num + 1)}1'].font = styles["bold_white_font"]
        worksheet[f'{get_excel_column_letter(column_num + 1)}1'].fill = styles["dark_blue_fill"]
        worksheet[f'{get_excel_column_letter(column_num + 1)}1'].alignment = styles["top_alignment"]

    for column_num, value in enumerate(data.iloc[-1]):
        worksheet[f'{get_excel_column_letter(column_num + 1)}{len(data) + 1}'] = value
        # Add formatting to the footer
        worksheet[f'{get_excel_column_letter(column_num + 1)}{len(data) + 1}'].font = styles["bold_black_font"]
        worksheet[f'{get_excel_column_letter(column_num + 1)}{len(data) + 1}'].fill = styles["yellow_fill"]
        worksheet[f'{get_excel_column_letter(column_num + 1)}{len(data) + 1}'].border = styles["thin_border"]
        worksheet[f'{get_excel_column_letter(column_num + 1)}{len(data) + 1}'].number_format = styles["numeric"]

    for column_num, width in enumerate(get_cols_widths(data)):
        worksheet.column_dimensions[get_excel_column_letter(column_num + 1)].width = width + 2

def columns_format(data: pd.DataFrame, worksheet: Worksheet) -> None:
    """
    Apply formatting to the columns of the worksheet.

    :param data: The exported dataframe
    :type data: pd.DataFrame
    :param worksheet: The worksheet to apply formatting to
    :type worksheet: Worksheet
    """
    number_columns: list[str] = ["Cliente", "Material", "Codigo Destinatario"]
    percentage_columns: list[str] = ["Bonificación", "Dto. Factura", "Dto. Adicional"]
    thousand_columns: list[str] = ["Valor neto", "PVP", "P. Crédito", "APORTE", "Importe NC"]
    list_columns: list[str] = number_columns + percentage_columns + thousand_columns
    for column_num, column_name in enumerate(data.columns):
        if column_name in list_columns:
            column_letter: str = get_excel_column_letter(column_num + 1)
            # Add formatting in each cell of the column
            for row in range(3, worksheet.max_row + 2):  # Skip the header row
                if column_name in number_columns:
                    worksheet[f'{column_letter}{row}'].number_format = styles["number"]
                elif column_name in percentage_columns:
                    worksheet[f'{column_letter}{row}'].number_format = styles["percentage"]
                elif column_name in thousand_columns:
                    worksheet[f'{column_letter}{row}'].number_format = styles["numeric"]

styles: dict[str, Any] = generate_styles()